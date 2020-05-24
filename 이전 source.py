import openpyxl
import tensorflow as tf
import numpy as np
import os
import matplotlib.pyplot as plt
import pygame

tf.set_random_seed(777)

seq_length=7
data_dim=6
output_dim=1
hidden_dim=5
num_studying=800
final_test=3861
num_stacked_layers=3

def MinMaxScaler(data_change,data):
    numerator = data_change - np.min(data, 0)
    denominator = np.max(data, 0) - np.min(data, 0)
    # noise term prevents the zero division
    return numerator / (denominator + 1e-7)

def reverse(new,data):
    min=np.min(data,0);
    max=np.max(data,0)
    return new*(max[5]-min[5]+1e-7)+min[5]

def lstm_cell():
    cell = tf.contrib.rnn.BasicLSTMCell(num_units=hidden_dim,state_is_tuple=True)
    return cell


#엑셀에서 데이터 가져오기
wb=openpyxl.load_workbook('timedata.xlsx')

ws=wb.active
x=[]
for r in ws.rows:
    x.append([r[1].value,r[2].value,r[3].value,r[4].value,r[5].value,r[6].value])
xy = MinMaxScaler(x,x)
#xy=x
#print(xy)


#데이터가공
dataX=[]
dataY=[]
for i in range(0,len(xy)-seq_length-final_test):
    _x=xy[i:i+seq_length]
    _y=xy[i+seq_length][data_dim-1]
    #print(_x,"->",_y)
    dataX.append(_x)
    dataY.append([_y])

#트레이닝 나누기
train_size=int(len(dataY)*0.8)
test_size=len(dataY)-train_size
trainX,testX=np.array(dataX[0:train_size]),np.array(dataX[train_size:len(dataY)])
trainY,testY=np.array(dataY[0:train_size]),np.array(dataY[train_size:len(dataY)])


#학습
X=tf.placeholder(tf.float32,[None,seq_length,data_dim])
Y=tf.placeholder(tf.float32,[None,1])

#cell=tf.contrib.rnn.BasicLSTMCell(num_units=hidden_dim,state_is_tuple=True)
multi_cells = tf.contrib.rnn.MultiRNNCell([lstm_cell() for _ in range(num_stacked_layers)],state_is_tuple=True)

#outputs,_states=tf.nn.dynamic_rnn(cell,X,dtype=tf.float32)
outputs,_states=tf.nn.dynamic_rnn(multi_cells,X,dtype=tf.float32)
Y_pred=tf.contrib.layers.fully_connected(outputs[:,-1],output_dim,activation_fn=None)

loss=tf.reduce_sum(tf.square(Y_pred-Y))

optimizer=tf.train.AdamOptimizer(0.01)
train=optimizer.minimize(loss)

targets=tf.placeholder(tf.float32,[None,1])
predictions=tf.placeholder(tf.float32,[None,1])
rmse=tf.sqrt(tf.reduce_mean(tf.square(targets-predictions)))

train_error_log=[]
test_error_log=[]

sess=tf.Session()
sess.run(tf.global_variables_initializer())

for i in range(num_studying):
    _,l=sess.run([train,loss],feed_dict={X: trainX, Y: trainY})

    if ((i+1)%10==0) or (i==num_studying-1):
        train_error=sess.run(rmse,feed_dict={targets:trainY,predictions:[[l]]})
        train_error_log.append(train_error)

        testPredict=sess.run(Y_pred,feed_dict={X: testX})
        rmse_val=sess.run(rmse,feed_dict={targets:testY,predictions:testPredict})
        test_error_log.append(rmse_val)

        print(i,l,format(rmse_val))


#test 실시
testdataX=[]
testdataY=[]
for i in range(len(xy)-final_test-1-seq_length,len(xy)-seq_length):
    _x=xy[i:i+seq_length]
    _y=xy[i+seq_length][data_dim-1]
    #print(_x,"->",_y)
    testdataX.append(_x)
    testdataY.append(_y)

testdataY=np.array(testdataY)
testdataY=reverse(testdataY,x)

print("Doing Test")
sum=0.0
result=[]
for i in range(0,len(testdataX)):
    if(i%50==0):
        print("Test : ",i,"/",len(testdataX))
    lastX=np.array([testdataX[i]])
    lastPredict=sess.run(Y_pred,feed_dict={X:lastX})
    lastPredict=reverse(lastPredict,x)
    aa=abs(testdataY[i]-lastPredict[0,0])/testdataY[i]*100
    sum=sum+aa
    #print(lastPredict[0,0],testdataY[i],100-aa)

    result.append(lastPredict[0,0])
print("average: ",100-sum/len(testdataX))



lastX=np.array([xy[len(xy)-seq_length:]])
#print(lastX)
lastPredict=sess.run(Y_pred,feed_dict={X:lastX})
#print("last day: ",reverse(lastPredict,x))
a=reverse(lastPredict,x)
data=a[0][0]

print("Making XLSX")
wr=openpyxl.load_workbook('score.xlsx')
wrr=wr.active

for i in range(0,len(testdataX)):
    if(i%50==0):
        print("Making : ",i,"/",len(testdataX))
    wrr.cell(row=i+1,column=1).value=result[i]
    #wrr.cell(row=i+1,column=1).value=1

wr.save('score.xlsx')
wr.close()



'''
plt.figure(1)
plt.plot(train_error_log)
plt.plot(test_error_log)
'''

plt.figure(2)
plt.plot(result)
plt.plot(testdataY)
plt.show()


"""demo용 initialize-start"""
pygame.init()
screen = pygame.display.set_mode((800,200))
clock = pygame.time.Clock()
run=True

# 이미지 로딩
# load() - 이미지파일을 읽어서 Surface 반환
background = pygame.image.load("graph.jpg")
moveimage = pygame.image.load("arrow.jpg")

pygame.display.set_caption("미세먼지 PM10")

h1 = pygame.font.SysFont('Sanserif', 36, False, False)
h2 = pygame.font.SysFont('Sanserif', 20, False, False)
text2 = h2.render('PM10', True,  [0,0,0])
"""demo용 initialize-end"""

datastr =str(data)
while run:
    for event in pygame.event.get():
        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_ESCAPE:
                run = False

    screen.fill(pygame.color.Color(255, 255, 255))
    pos = 0
    if data <=30:
        str = "good"
        pos=190*(data/30)
    elif data <=80:
        str = "normal"
        pos = 200 * ((data-30) / 50) + 190
    elif data <= 150:
        str = "bad"
        pos = 197 * ((data - 80) / 70) + 390
    else:
        str ="very bad"
        pos= 213*((data - 150) / 300) + 587


    text = h1.render(str, True, [0, 0, 0])
    text3 = h2.render(datastr, True, [0, 0, 0])
    #Surface.blit(Surface객체, (x좌표, y좌표))
    screen.blit(text, (5,0))
    screen.blit(text2, (8, 25))
    """
    좋음(0~30) : 0~190
    보통(30~80) : 190~390
    나쁨(80~150) : 390~587
    매우나쁨(150~) : 587~800
    """

    screen.blit(text3, (pos, 40))
    screen.blit(moveimage, (pos, 55))
    screen.blit(background,(0,100))

    pygame.display.flip()

#    clock.tick(60)

pygame.quit()
